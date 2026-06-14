"""
Generate template_gsheet.xlsx untuk diupload ke Google Sheets.

Worksheet:
  PETUNJUK   : panduan upload + flowchart alur (visual)
  KALKULATOR : kalkulator self-contained dengan formula
               (user edit absorbansi -> %inhibisi/IC50 auto-hitung)
  DPPH       : long-format untuk aplikasi Streamlit + 5 contoh
  UAE        : header + 1 contoh
  TPC        : header + 1 contoh
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[1]))

from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Font,
    PatternFill,
    Border,
    Side,
    NamedStyle,
)
from openpyxl.utils import get_column_letter

from utils.sheets import DPPH_COLUMNS, UAE_COLUMNS, TPC_COLUMNS
from utils.calculations import process_experiment

OUT = Path(__file__).resolve().parents[2] / "template_gsheet.xlsx"

# ============================================================
# Color palette (hijau alami sesuai topik daun salam)
# ============================================================
GREEN_DARK = "1B5E20"
GREEN_MAIN = "2E7D32"
GREEN_MID = "4CAF50"
GREEN_LIGHT = "C8E6C9"
GREEN_PALE = "F1F8E9"
ORANGE_MAIN = "EF6C00"
ORANGE_LIGHT = "FFE0B2"
BLUE_MAIN = "1565C0"
BLUE_LIGHT = "BBDEFB"
GREY_BORDER = "B0BEC5"
GREY_TEXT = "455A64"
YELLOW_LIGHT = "FFF59D"
YELLOW_MID = "FFD54F"

# ============================================================
# Style primitives
# ============================================================
THIN = Side(border_style="thin", color=GREY_BORDER)
THICK = Side(border_style="medium", color=GREEN_DARK)
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
BORDER_BOX = Border(left=THICK, right=THICK, top=THICK, bottom=THICK)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def style_cell(cell, font=None, bg=None, border=None, align=None, num_fmt=None):
    if font:
        cell.font = font
    if bg:
        cell.fill = fill(bg)
    if border:
        cell.border = border
    if align:
        cell.alignment = align
    if num_fmt:
        cell.number_format = num_fmt


def header_block(ws, row: int, cols: int, text: str, *, bg=GREEN_MAIN, fg="FFFFFF"):
    """Header bold row dengan merged cells."""
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(
        start_row=row,
        start_column=1,
        end_row=row,
        end_column=cols,
    )
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        style_cell(
            cell,
            font=Font(bold=True, color=fg, size=12),
            bg=bg,
            align=CENTER,
        )
    ws.row_dimensions[row].height = 26


def style_table_header(ws, row: int, n_cols: int):
    """Header tabel data DPPH/UAE/TPC."""
    for c in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=c)
        style_cell(
            cell,
            font=Font(bold=True, color="FFFFFF", size=11),
            bg=GREEN_MAIN,
            align=Alignment(horizontal="center", vertical="center", wrap_text=True),
            border=BORDER_THIN,
        )
    ws.row_dimensions[row].height = 32


def autofit(ws, min_w: int = 10, max_w: int = 28):
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_col=col_idx, max_col=col_idx):
            for cell in row:
                v = "" if cell.value is None else str(cell.value)
                for line in v.split("\n"):
                    if len(line) > max_len:
                        max_len = len(line)
        ws.column_dimensions[letter].width = max(min_w, min(max_len + 2, max_w))


def write_table(ws, start_row: int, df: pd.DataFrame):
    """
    Tulis dataframe ke worksheet.

    NaN values dirender sebagai '-' (strip) supaya cell tidak terlihat
    kosong/forgot-to-fill. Aplikasi & pandas tetap baca '-' sebagai NaN
    via pd.to_numeric(errors='coerce').
    """
    import math

    for j, col in enumerate(df.columns, start=1):
        ws.cell(row=start_row, column=j, value=col)
    style_table_header(ws, start_row, len(df.columns))
    for i, row in enumerate(df.itertuples(index=False), start=start_row + 1):
        for j, val in enumerate(row, start=1):
            # Render NaN sebagai strip "-"
            is_nan = val is None or (isinstance(val, float) and math.isnan(val))
            if is_nan:
                cell = ws.cell(row=i, column=j, value="-")
                style_cell(
                    cell,
                    font=Font(color=GREY_TEXT, size=10),
                    bg="EEEEEE",
                    border=BORDER_THIN,
                    align=Alignment(horizontal="center", vertical="center"),
                )
                continue

            cell = ws.cell(row=i, column=j, value=val)
            cell.border = BORDER_THIN
            if isinstance(val, float):
                cell.number_format = "0.0000"
            cell.alignment = Alignment(horizontal="center", vertical="center")


# ============================================================
# Tab 1: PETUNJUK — visual flowchart
# ============================================================
def _draw_box(
    ws,
    top_row: int,
    left_col: int,
    height: int,
    width: int,
    text: str,
    *,
    bg: str,
    border_color: str = GREEN_DARK,
    text_color: str = "FFFFFF",
    bold: bool = True,
    size: int = 11,
):
    """Gambar box (merged cells) dengan teks di tengah."""
    ws.merge_cells(
        start_row=top_row,
        start_column=left_col,
        end_row=top_row + height - 1,
        end_column=left_col + width - 1,
    )
    cell = ws.cell(row=top_row, column=left_col, value=text)
    border = Border(
        left=Side("medium", color=border_color),
        right=Side("medium", color=border_color),
        top=Side("medium", color=border_color),
        bottom=Side("medium", color=border_color),
    )
    style_cell(
        cell,
        font=Font(bold=bold, color=text_color, size=size),
        bg=bg,
        border=border,
        align=CENTER,
    )
    # Apply border to all cells in merged range
    for r in range(top_row, top_row + height):
        for c in range(left_col, left_col + width):
            cc = ws.cell(row=r, column=c)
            if cc.border != border:
                # Border merging: pakai border consistent
                cc.border = border
                cc.fill = fill(bg)


def _draw_arrow(
    ws, row: int, col: int, *, direction: str = "down", color: str = GREEN_DARK
):
    """Gambar panah ASCII satu cell."""
    arrow_map = {
        "down": "▼",
        "up": "▲",
        "right": "▶",
        "left": "◀",
    }
    cell = ws.cell(row=row, column=col, value=arrow_map[direction])
    style_cell(
        cell,
        font=Font(bold=True, color=color, size=18),
        align=CENTER,
    )


def build_petunjuk(ws):
    # ---------- Title ----------
    ws.cell(
        row=1, column=1, value="TEMPLATE GOOGLE SHEETS — PLATFORM EVALUASI ANTIOKSIDAN"
    )
    ws.merge_cells("A1:L1")
    style_cell(
        ws.cell(row=1, column=1),
        font=Font(bold=True, color=GREEN_DARK, size=18),
        align=CENTER,
    )
    ws.row_dimensions[1].height = 32

    ws.cell(
        row=2,
        column=1,
        value="Skema database untuk modul DPPH (uji antioksidan), UAE (parameter ekstraksi), dan TPC (total fenolik)",
    )
    ws.merge_cells("A2:L2")
    style_cell(
        ws.cell(row=2, column=1),
        font=Font(italic=True, color=GREY_TEXT, size=11),
        align=CENTER,
    )
    ws.row_dimensions[2].height = 22

    # ---------- Section: Daftar Tab ----------
    header_block(ws, 4, 12, "ISI TEMPLATE")

    tab_info = [
        ("PETUNJUK", "Halaman ini — panduan + diagram alur", GREEN_PALE, GREEN_DARK),
        (
            "KALKULATOR",
            "Kalkulator interaktif: input absorbansi -> auto-hitung",
            YELLOW_LIGHT,
            ORANGE_MAIN,
        ),
        (
            "DPPH",
            "Database long-format untuk aplikasi Streamlit",
            GREEN_LIGHT,
            GREEN_DARK,
        ),
        ("UAE", "Parameter ekstraksi Ultrasound-Assisted", BLUE_LIGHT, BLUE_MAIN),
        ("TPC", "Total Phenolic Content (Folin-Ciocalteu)", ORANGE_LIGHT, ORANGE_MAIN),
    ]
    row = 5
    for tab_name, desc, bg_c, text_c in tab_info:
        ws.cell(row=row, column=1, value=tab_name)
        style_cell(
            ws.cell(row=row, column=1),
            font=Font(bold=True, color=text_c, size=11),
            bg=bg_c,
            align=CENTER,
            border=BORDER_THIN,
        )
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        ws.cell(row=row, column=2, value=desc)
        for c in range(2, 13):
            style_cell(
                ws.cell(row=row, column=c),
                bg=bg_c,
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.row_dimensions[row].height = 22
        row += 1

    # ---------- Section: Cara Upload ----------
    row += 1
    header_block(ws, row, 12, "CARA UPLOAD KE GOOGLE SHEETS")
    row += 1
    upload_steps = [
        "1. Buka https://sheets.google.com lalu klik 'Blank' (spreadsheet kosong)",
        "2. Menu File > Import > Upload > pilih file template_gsheet.xlsx ini",
        "3. Pilih opsi 'Replace spreadsheet' lalu klik 'Import data'",
        "4. Tunggu proses upload (sekitar 5-10 detik)",
        "5. Selesai — 5 tab muncul: PETUNJUK, KALKULATOR, DPPH, UAE, TPC",
        "6. Salin URL spreadsheet untuk dipasang di secrets.toml aplikasi",
    ]
    for step in upload_steps:
        ws.cell(row=row, column=1, value=step)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
        style_cell(
            ws.cell(row=row, column=1),
            font=Font(size=11),
            align=LEFT,
            border=BORDER_THIN,
        )
        ws.row_dimensions[row].height = 22
        row += 1

    # ---------- Section: Flowchart Alur Data ----------
    row += 2
    header_block(ws, row, 12, "ALUR KERJA RISET")
    row += 2

    # Flowchart visual menggunakan box + arrow
    # Layout 12 kolom, lebar tiap box = 4 kolom
    #
    #  Layer 1:   [ Lab: Spektrofotometer ]
    #                       |
    #  Layer 2:   [ Input data ke Streamlit / Edit di tab KALKULATOR ]
    #                       |
    #  Layer 3:   [ Auto-hitung: %inhibisi, IC50, regresi ]
    #                       |
    #  Layer 4:   [ Tersimpan di tab DPPH ]
    #                       |
    #               +-------+-------+-------+
    #  Layer 5:  [Visualisasi] [ANOVA] [PDF Report]

    # Layer 1
    _draw_box(
        ws,
        row,
        5,
        height=2,
        width=4,
        text="LAB\nPengukuran absorbansi DPPH\nspektrofotometer 517 nm",
        bg=ORANGE_MAIN,
        text_color="FFFFFF",
    )
    row += 2
    _draw_arrow(ws, row, 6)
    _draw_arrow(ws, row, 7)
    row += 1

    # Layer 2
    _draw_box(
        ws,
        row,
        4,
        height=2,
        width=6,
        text="INPUT DATA\nLewat aplikasi Streamlit (page Input DPPH)\nATAU edit langsung di tab KALKULATOR",
        bg=BLUE_MAIN,
        text_color="FFFFFF",
    )
    row += 2
    _draw_arrow(ws, row, 6)
    _draw_arrow(ws, row, 7)
    row += 1

    # Layer 3
    _draw_box(
        ws,
        row,
        4,
        height=2,
        width=6,
        text="PERHITUNGAN OTOMATIS\n% inhibisi (pairwise), mean, SD,\nregresi linear, IC50, kategori",
        bg=GREEN_MAIN,
        text_color="FFFFFF",
    )
    row += 2
    _draw_arrow(ws, row, 6)
    _draw_arrow(ws, row, 7)
    row += 1

    # Layer 4
    _draw_box(
        ws,
        row,
        4,
        height=2,
        width=6,
        text="DATABASE\nWorksheet DPPH (long-format)\nsource of truth",
        bg=GREEN_DARK,
        text_color="FFFFFF",
        size=12,
    )
    row += 2

    # Branch arrows
    branch_row = row
    _draw_arrow(ws, branch_row, 3)
    _draw_arrow(ws, branch_row, 6)
    _draw_arrow(ws, branch_row, 7)
    _draw_arrow(ws, branch_row, 11)
    row += 1

    # Layer 5: 3 outputs
    _draw_box(
        ws,
        row,
        1,
        height=2,
        width=4,
        text="VISUALISASI\nKurva regresi,\nbar chart, line chart",
        bg=YELLOW_MID,
        text_color=GREEN_DARK,
    )
    _draw_box(
        ws,
        row,
        5,
        height=2,
        width=4,
        text="ANALISIS ANOVA\nUji beda nyata\n+ Tukey HSD",
        bg=YELLOW_MID,
        text_color=GREEN_DARK,
    )
    _draw_box(
        ws,
        row,
        9,
        height=2,
        width=4,
        text="EXPORT PDF\nLaporan per percobaan\nuntuk lampiran tesis",
        bg=YELLOW_MID,
        text_color=GREEN_DARK,
    )
    row += 3

    # ---------- Section: Aturan Struktur ----------
    header_block(ws, row, 12, "ATURAN STRUKTUR DATA")
    row += 1
    rules = [
        ("WAJIB", "Baris pertama setiap tab = header. JANGAN diubah/dihapus/digeser."),
        ("WAJIB", "Nama kolom case-sensitive. Harus persis sama dengan template ini."),
        ("WAJIB", "experiment_id unik per percobaan."),
        ("WAJIB", "Konsentrasi 0 ppm = blanko (DPPH + pelarut tanpa sampel)."),
        (
            "WAJIB",
            "1 percobaan DPPH = 6 baris (konsentrasi 0, 20, 40, 60, 80, 100 ppm).",
        ),
        ("WAJIB", "Pemisah desimal pakai titik (.), bukan koma."),
        ("OPSIONAL", "Tab KALKULATOR untuk quick check tanpa buka aplikasi."),
        ("OPSIONAL", "Tab PETUNJUK boleh dihapus setelah dibaca."),
    ]
    for tag, rule in rules:
        # Tag column (bold colored)
        ws.cell(row=row, column=1, value=tag)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        tag_color = ORANGE_MAIN if tag == "WAJIB" else BLUE_MAIN
        for c in range(1, 3):
            style_cell(
                ws.cell(row=row, column=c),
                font=Font(bold=True, color="FFFFFF", size=10),
                bg=tag_color,
                align=CENTER,
                border=BORDER_THIN,
            )
        # Rule text
        ws.cell(row=row, column=3, value=rule)
        ws.merge_cells(start_row=row, start_column=3, end_row=row, end_column=12)
        for c in range(3, 13):
            style_cell(
                ws.cell(row=row, column=c),
                font=Font(size=10),
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.row_dimensions[row].height = 20
        row += 1

    # ---------- Section: Rumus ----------
    row += 1
    header_block(ws, row, 12, "RUMUS PERHITUNGAN")
    row += 1

    formulas = [
        (
            "% Inhibisi (pairwise)",
            "Inhibisi_i  =  ((Abs_blanko_i  −  Abs_sampel_i)  /  Abs_blanko_i)  ×  100",
        ),
        (
            "Regresi linear",
            "y  =  a · x  +  b      (x = konsentrasi ppm,  y = % inhibisi rata-rata)",
        ),
        (
            "IC50",
            "IC50  =  (50  −  b)  /  a      satuan: ppm",
        ),
        (
            "Kategori (Molyneux 2004)",
            "<50 Sangat kuat  |  50-100 Kuat  |  100-150 Sedang  |  150-200 Lemah  |  >200 Sangat lemah",
        ),
    ]
    for label, formula in formulas:
        ws.cell(row=row, column=1, value=label)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
        for c in range(1, 5):
            style_cell(
                ws.cell(row=row, column=c),
                font=Font(bold=True, color=GREEN_DARK, size=10),
                bg=GREEN_PALE,
                align=LEFT,
                border=BORDER_THIN,
            )
        # Important: pakai prefix space supaya '=' di awal string tidak
        # dianggap formula oleh Excel/Google Sheets. Lalu strip via formatting.
        # Kita pakai pattern "Label = expression" jadi "=" di tengah, bukan awal.
        cell = ws.cell(row=row, column=5)
        cell.value = formula  # sudah aman, tidak diawali '='
        ws.merge_cells(start_row=row, start_column=5, end_row=row, end_column=12)
        for c in range(5, 13):
            style_cell(
                ws.cell(row=row, column=c),
                font=Font(name="Consolas", size=10),
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.row_dimensions[row].height = 22
        row += 1

    # ---------- Footer ----------
    row += 1
    ws.cell(
        row=row,
        column=1,
        value="Platform Digital Evaluasi Antioksidan Daun Salam — generated automatically",
    )
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=12)
    style_cell(
        ws.cell(row=row, column=1),
        font=Font(italic=True, color=GREY_TEXT, size=9),
        align=CENTER,
    )

    # Column widths (12 kolom equal)
    for col_idx in range(1, 13):
        ws.column_dimensions[get_column_letter(col_idx)].width = 13

    # Hide gridlines untuk tampilan clean (di Google Sheets juga ke-respect)
    ws.sheet_view.showGridLines = False


# ============================================================
# Tab 2: KALKULATOR — self-contained dengan formula
# ============================================================
def build_kalkulator(ws):
    """
    Layout KALKULATOR:

    Row 1-2  : Title
    Row 4-7  : Section "1. METADATA" — text input (sampel, tanggal, waktu inkubasi)
    Row 9-10 : Section header "2. INPUT ABSORBANSI"
    Row 11   : Header tabel (Konsentrasi, Abs1, Abs2, Abs3, Mean, Inhib1, Inhib2, Inhib3, Mean, SD)
    Row 12-17: Data 6 row (konsentrasi 0, 20, 40, 60, 80, 100)
    Row 19-20: Section "3. HASIL"
    Row 21-26: Output cells (slope, intercept, R², IC50, kategori, persamaan)
    Row 28+  : Petunjuk pemakaian
    """
    # ---------- Title ----------
    ws.cell(row=1, column=1, value="KALKULATOR DPPH — UJI AKTIVITAS ANTIOKSIDAN")
    ws.merge_cells("A1:J1")
    style_cell(
        ws.cell(row=1, column=1),
        font=Font(bold=True, color=GREEN_DARK, size=16),
        align=CENTER,
    )
    ws.row_dimensions[1].height = 30

    ws.cell(
        row=2,
        column=1,
        value="Edit absorbansi pada cell warna kuning. Hasil otomatis terhitung di cell warna hijau.",
    )
    ws.merge_cells("A2:J2")
    style_cell(
        ws.cell(row=2, column=1),
        font=Font(italic=True, color=GREY_TEXT, size=10),
        align=CENTER,
    )

    # ---------- Section 1: Metadata ----------
    header_block(ws, 4, 10, "1. METADATA PERCOBAAN", bg=BLUE_MAIN)

    metadata = [
        ("Tanggal", "2026-01-15"),
        ("Sampel", "Ekstrak Etanol Daun Salam"),
        ("Metode ekstraksi", "UAE"),
        ("Waktu inkubasi (menit)", 5),
    ]
    for i, (label, default_val) in enumerate(metadata):
        r = 5 + i
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
        for c in range(1, 4):
            style_cell(
                ws.cell(row=r, column=c),
                font=Font(bold=True, color=GREEN_DARK, size=11),
                bg=GREEN_PALE,
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.cell(row=r, column=4, value=default_val)
        ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=10)
        for c in range(4, 11):
            style_cell(
                ws.cell(row=r, column=c),
                font=Font(size=11),
                bg=YELLOW_LIGHT,  # editable
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.row_dimensions[r].height = 20

    # ---------- Section 2: Input Absorbansi ----------
    header_block(ws, 10, 10, "2. INPUT ABSORBANSI (3 REPLIKASI)", bg=ORANGE_MAIN)

    # Header tabel
    headers = [
        "Konsentrasi\n(ppm)",
        "Abs 1",
        "Abs 2",
        "Abs 3",
        "Abs Mean",
        "Inhibisi 1\n(%)",
        "Inhibisi 2\n(%)",
        "Inhibisi 3\n(%)",
        "Inhib Mean\n(%)",
        "SD",
    ]
    for col, h in enumerate(headers, start=1):
        cell = ws.cell(row=11, column=col, value=h)
        style_cell(
            cell,
            font=Font(bold=True, color="FFFFFF", size=10),
            bg=GREEN_MAIN,
            align=CENTER,
            border=BORDER_THIN,
        )
    ws.row_dimensions[11].height = 36

    # Data input (default values dari template DPPH 5 menit)
    data_input = [
        (0, 0.667, 0.669, 0.671),
        (20, 0.469, 0.464, 0.465),
        (40, 0.399, 0.398, 0.392),
        (60, 0.268, 0.267, 0.267),
        (80, 0.141, 0.142, 0.142),
        (100, 0.051, 0.053, 0.055),
    ]

    # Row 12-17: tabel data
    for i, (konsentrasi, a1, a2, a3) in enumerate(data_input):
        r = 12 + i
        # Kolom A: konsentrasi (editable kuning)
        ws.cell(row=r, column=1, value=konsentrasi)
        style_cell(
            ws.cell(row=r, column=1),
            font=Font(bold=True, size=11),
            bg=YELLOW_LIGHT,
            align=CENTER,
            border=BORDER_THIN,
            num_fmt="0",
        )
        # Kolom B-D: abs 1/2/3 (editable kuning)
        for j, val in enumerate([a1, a2, a3], start=2):
            ws.cell(row=r, column=j, value=val)
            style_cell(
                ws.cell(row=r, column=j),
                font=Font(size=11),
                bg=YELLOW_LIGHT,
                align=CENTER,
                border=BORDER_THIN,
                num_fmt="0.0000",
            )
        # Kolom E: abs mean (formula)
        ws.cell(row=r, column=5, value=f"=AVERAGE(B{r}:D{r})")
        style_cell(
            ws.cell(row=r, column=5),
            font=Font(size=11, color=GREEN_DARK),
            bg=GREEN_PALE,
            align=CENTER,
            border=BORDER_THIN,
            num_fmt="0.0000",
        )
        # Kolom F-H: %inhibisi 1/2/3 (formula pairwise, blanko di row 12)
        if r == 12:  # blanko: kosong
            for c in range(6, 9):
                cell = ws.cell(row=r, column=c, value="")
                style_cell(
                    cell,
                    bg="EEEEEE",
                    align=CENTER,
                    border=BORDER_THIN,
                )
            ws.cell(row=r, column=9, value="")
            style_cell(
                ws.cell(row=r, column=9), bg="EEEEEE", align=CENTER, border=BORDER_THIN
            )
            ws.cell(row=r, column=10, value="")
            style_cell(
                ws.cell(row=r, column=10), bg="EEEEEE", align=CENTER, border=BORDER_THIN
            )
        else:
            # %inhib 1 = ((B$12 - B<r>) / B$12) * 100
            for c, ref in zip([6, 7, 8], ["B", "C", "D"]):
                formula = f'=IFERROR((({ref}$12-{ref}{r})/{ref}$12)*100, "")'
                ws.cell(row=r, column=c, value=formula)
                style_cell(
                    ws.cell(row=r, column=c),
                    font=Font(size=11, color=GREEN_DARK),
                    bg=GREEN_PALE,
                    align=CENTER,
                    border=BORDER_THIN,
                    num_fmt="0.0000",
                )
            # Inhib mean = AVERAGE(F<r>:H<r>)
            ws.cell(row=r, column=9, value=f'=IFERROR(AVERAGE(F{r}:H{r}), "")')
            style_cell(
                ws.cell(row=r, column=9),
                font=Font(bold=True, size=11, color=GREEN_DARK),
                bg=GREEN_LIGHT,
                align=CENTER,
                border=BORDER_THIN,
                num_fmt="0.0000",
            )
            # SD = STDEV(F<r>:H<r>)  (sample SD, ddof=1)
            ws.cell(row=r, column=10, value=f'=IFERROR(STDEV(F{r}:H{r}), "")')
            style_cell(
                ws.cell(row=r, column=10),
                font=Font(size=11, color=GREEN_DARK),
                bg=GREEN_PALE,
                align=CENTER,
                border=BORDER_THIN,
                num_fmt="0.0000",
            )

    # ---------- Section 3: Hasil Regresi & IC50 ----------
    header_block(ws, 19, 10, "3. HASIL REGRESI LINEAR & IC50", bg=GREEN_MAIN)

    # Regresi: pakai SLOPE / INTERCEPT / RSQ pada data konsen vs inhib_mean
    # Range data: x = A13:A17 (konsentrasi 20-100), y = I13:I17 (inhib mean)
    results = [
        ("Slope (a)", "=SLOPE(I13:I17, A13:A17)", "0.0000", False),
        ("Intercept (b)", "=INTERCEPT(I13:I17, A13:A17)", "0.0000", False),
        ("R-squared", "=RSQ(I13:I17, A13:A17)", "0.0000", False),
        (
            "Persamaan",
            '=("y = " & TEXT(SLOPE(I13:I17, A13:A17), "0.0000") '
            '& " · x + " & TEXT(INTERCEPT(I13:I17, A13:A17), "0.0000"))',
            "@",
            False,
        ),
        (
            "IC50 (ppm)",
            "=(50 - INTERCEPT(I13:I17, A13:A17)) / SLOPE(I13:I17, A13:A17)",
            "0.0000",
            True,
        ),
        (
            "Kategori",
            (
                "=IFS("
                'F24<50, "Sangat kuat", '
                'F24<100, "Kuat", '
                'F24<150, "Sedang", '
                'F24<200, "Lemah", '
                'TRUE, "Sangat lemah")'
            ),
            "@",
            True,
        ),
    ]
    for i, (label, formula, fmt, highlight) in enumerate(results):
        r = 20 + i
        ws.cell(row=r, column=1, value=label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)
        for c in range(1, 6):
            style_cell(
                ws.cell(row=r, column=c),
                font=Font(bold=True, color=GREEN_DARK, size=11),
                bg=GREEN_PALE,
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.cell(row=r, column=6, value=formula)
        ws.merge_cells(start_row=r, start_column=6, end_row=r, end_column=10)
        bg_result = YELLOW_MID if highlight else GREEN_LIGHT
        for c in range(6, 11):
            style_cell(
                ws.cell(row=r, column=c),
                font=Font(
                    bold=highlight, color=GREEN_DARK, size=12 if highlight else 11
                ),
                bg=bg_result,
                align=CENTER,
                border=BORDER_THIN,
                num_fmt=fmt if c == 6 else None,
            )
        ws.row_dimensions[r].height = 24 if highlight else 20

    # ---------- Section 4: Petunjuk ----------
    header_block(ws, 28, 10, "PETUNJUK PEMAKAIAN", bg=GREY_TEXT)

    helps = [
        (
            "KUNING",
            "Cell yang BOLEH diedit (input absorbansi & metadata).",
            YELLOW_LIGHT,
        ),
        ("HIJAU", "Cell formula — JANGAN diedit, otomatis ke-update.", GREEN_LIGHT),
        ("ABU", "Blanko (konsentrasi 0 ppm) — % inhibisi tidak dihitung.", "EEEEEE"),
        ("KUNING TUA", "Hasil utama: IC50 dan Kategori antioksidan.", YELLOW_MID),
    ]
    for i, (color_label, desc, bg_c) in enumerate(helps):
        r = 29 + i
        ws.cell(row=r, column=1, value=color_label)
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=2)
        for c in range(1, 3):
            style_cell(
                ws.cell(row=r, column=c),
                font=Font(bold=True, color=GREEN_DARK, size=10),
                bg=bg_c,
                align=CENTER,
                border=BORDER_THIN,
            )
        ws.cell(row=r, column=3, value=desc)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=10)
        for c in range(3, 11):
            style_cell(
                ws.cell(row=r, column=c),
                font=Font(size=10),
                align=LEFT,
                border=BORDER_THIN,
            )
        ws.row_dimensions[r].height = 20

    # ---------- Tips ----------
    r = 34
    ws.cell(
        row=r,
        column=1,
        value=(
            "TIPS: Hasil dari kalkulator ini bisa di-copy manual ke tab DPPH "
            "(sebagai 1 percobaan = 6 baris). "
            "Atau biarkan aplikasi Streamlit yang menulis ke tab DPPH otomatis."
        ),
    )
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=10)
    style_cell(
        ws.cell(row=r, column=1),
        font=Font(italic=True, color=GREY_TEXT, size=10),
        align=LEFT,
    )
    ws.row_dimensions[r].height = 36

    # Column widths
    widths = [12, 10, 10, 10, 11, 11, 11, 11, 11, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A11"  # freeze metadata + section header


# ============================================================
# Tab 3: DPPH (long-format, untuk aplikasi)
# ============================================================
def build_dpph_examples() -> pd.DataFrame:
    """5 percobaan contoh dengan variasi waktu inkubasi 5/6/7/8/9 menit."""
    examples = []

    sets = {
        5: pd.DataFrame(
            {
                "konsentrasi": [0, 20, 40, 60, 80, 100],
                "abs_1": [0.667, 0.469, 0.399, 0.268, 0.141, 0.051],
                "abs_2": [0.669, 0.464, 0.398, 0.267, 0.142, 0.053],
                "abs_3": [0.671, 0.465, 0.392, 0.267, 0.142, 0.055],
            }
        ),
        6: pd.DataFrame(
            {
                "konsentrasi": [0, 20, 40, 60, 80, 100],
                "abs_1": [0.679, 0.489, 0.428, 0.379, 0.231, 0.112],
                "abs_2": [0.677, 0.492, 0.431, 0.371, 0.233, 0.115],
                "abs_3": [0.675, 0.495, 0.423, 0.373, 0.236, 0.117],
            }
        ),
        7: pd.DataFrame(
            {
                "konsentrasi": [0, 20, 40, 60, 80, 100],
                "abs_1": [0.672, 0.495, 0.439, 0.396, 0.285, 0.155],
                "abs_2": [0.678, 0.497, 0.447, 0.387, 0.297, 0.157],
                "abs_3": [0.678, 0.491, 0.449, 0.397, 0.294, 0.159],
            }
        ),
        8: pd.DataFrame(
            {
                "konsentrasi": [0, 20, 40, 60, 80, 100],
                "abs_1": [0.677, 0.516, 0.469, 0.418, 0.311, 0.191],
                "abs_2": [0.679, 0.524, 0.458, 0.407, 0.312, 0.193],
                "abs_3": [0.671, 0.515, 0.462, 0.417, 0.302, 0.195],
            }
        ),
        9: pd.DataFrame(
            {
                "konsentrasi": [0, 20, 40, 60, 80, 100],
                "abs_1": [0.679, 0.539, 0.478, 0.419, 0.331, 0.212],
                "abs_2": [0.678, 0.532, 0.461, 0.411, 0.333, 0.215],
                "abs_3": [0.675, 0.535, 0.473, 0.423, 0.336, 0.217],
            }
        ),
    }

    base_date = "2026-01-15"
    for waktu, df_in in sets.items():
        df_proc, reg = process_experiment(df_in, blanko_method="pairwise")
        exp_id = f"DPPH-{base_date.replace('-', '')}-{waktu:02d}MIN"
        for _, r in df_proc.iterrows():
            examples.append(
                {
                    "experiment_id": exp_id,
                    "tanggal": base_date,
                    "sampel": "Ekstrak Etanol Daun Salam",
                    "metode_ekstraksi": "UAE",
                    "waktu_inkubasi_menit": waktu,
                    "konsentrasi": r["konsentrasi"],
                    "abs_1": r["abs_1"],
                    "abs_2": r["abs_2"],
                    "abs_3": r["abs_3"],
                    "abs_mean": r["abs_mean"],
                    "inhib_1": r.get("inhib_1"),
                    "inhib_2": r.get("inhib_2"),
                    "inhib_3": r.get("inhib_3"),
                    "inhib_mean": r.get("inhib_mean"),
                    "inhib_sd": r.get("inhib_sd"),
                    "ic50_ppm": reg.ic50,
                    "r_squared": reg.r_squared,
                    "persamaan_regresi": reg.equation,
                    "catatan": "",
                    "created_at": datetime.now().isoformat(timespec="seconds"),
                }
            )
    return pd.DataFrame(examples, columns=DPPH_COLUMNS)


def build_dpph(ws):
    df = build_dpph_examples()
    write_table(ws, 1, df)
    autofit(ws)
    ws.freeze_panes = "A2"


# ============================================================
# Tab 4: UAE
# ============================================================
def build_uae(ws):
    rows = [
        {
            "tanggal": "2026-01-10",
            "kode_sampel": "UAE-001",
            "massa_simplisia_g": 10.0,
            "volume_pelarut_ml": 100.0,
            "jenis_pelarut": "Etanol 70%",
            "rasio_pelarut": "1:10",
            "amplitudo_persen": 60,
            "frekuensi_khz": 20.0,
            "suhu_c": 40.0,
            "waktu_sonikasi_menit": 30,
            "siklus_on_off": "5/3",
            "massa_ekstrak_g": 1.85,
            "rendemen_persen": 18.50,
            "catatan": "",
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
    ]
    df = pd.DataFrame(rows, columns=UAE_COLUMNS)
    write_table(ws, 1, df)
    autofit(ws)
    ws.freeze_panes = "A2"


# ============================================================
# Tab 5: TPC
# ============================================================
def build_tpc(ws):
    rows = [
        {
            "tanggal": "2026-01-12",
            "kode_sampel": "TPC-001",
            "konsentrasi_ppm": 100.0,
            "abs_1": 0.512,
            "abs_2": 0.518,
            "abs_3": 0.515,
            "abs_mean": 0.515,
            "tpc_mg_gae_per_g": 45.32,
            "catatan": "",
            "created_at": datetime.now().isoformat(timespec="seconds"),
        }
    ]
    df = pd.DataFrame(rows, columns=TPC_COLUMNS)
    write_table(ws, 1, df)
    autofit(ws)
    ws.freeze_panes = "A2"


# ============================================================
def main():
    wb = Workbook()

    ws_p = wb.active
    ws_p.title = "PETUNJUK"
    build_petunjuk(ws_p)

    ws_k = wb.create_sheet("KALKULATOR")
    build_kalkulator(ws_k)

    ws_d = wb.create_sheet("DPPH")
    build_dpph(ws_d)

    ws_u = wb.create_sheet("UAE")
    build_uae(ws_u)

    ws_t = wb.create_sheet("TPC")
    build_tpc(ws_t)

    wb.save(OUT)
    print(f"Wrote: {OUT}")
    print(f"Sheets: PETUNJUK, KALKULATOR, DPPH, UAE, TPC")


if __name__ == "__main__":
    main()
